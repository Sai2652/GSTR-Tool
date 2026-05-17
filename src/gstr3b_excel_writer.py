"""
GSTR-3B Excel output writer.

Produces a polished multi-sheet workbook for the client / accountant.
Sheets:
    1. Summary           - one-page snapshot
    2. ITC Details       - GSTR-2B breakdown by category
    3. Output Tax        - Table 3.1 figures
    4. ITC Computation   - Table 4 (available - reversal = net)
    5. Tax Setoff        - step-by-step set-off trail
    6. Cash Payable      - final cash by tax head
    7. Credit Ledger     - opening + ITC - used = closing

Public API:
    write_gstr3b_excel(out_path, firm, period, gstr2b, computation)
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter


# ---- Brand palette --------------------------------------------------------

NIGHT_BLUE = "FF001441"
PURPLE = "FF6958C2"
PURPLE_LIGHT = "FFE9E4F5"
PALE_LAVENDER = "FFF7F5FC"
WHITE = "FFFFFFFF"
GREEN = "FF15803D"
RED = "FFB91C1C"
GREY_MUTED = "FF6B7280"
GREY_BORDER = "FFD0D0D8"


# ---- Reusable styles -----------------------------------------------------

def _font(size=11, bold=False, color="FF1A1A2E", italic=False) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _thin_border() -> Border:
    side = Side(style="thin", color=GREY_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---- Layout helpers ------------------------------------------------------

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title_row(ws, row, text, span_cols, fill=NIGHT_BLUE, color=WHITE, size=14, height=28):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    cell = ws.cell(row, 1, text)
    cell.font = _font(size=size, bold=True, color=color)
    cell.alignment = _center()
    cell.fill = _fill(fill)
    ws.row_dimensions[row].height = height


def _section_row(ws, row, text, span_cols, fill=PURPLE_LIGHT, color="FF1A1A2E"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    cell = ws.cell(row, 1, text)
    cell.font = _font(size=11, bold=True, color=color)
    cell.alignment = _left()
    cell.fill = _fill(fill)
    ws.row_dimensions[row].height = 22


def _table_header(ws, row, headers, fill=PURPLE, color=WHITE):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        c.font = _font(size=10, bold=True, color=color)
        c.alignment = _center()
        c.fill = _fill(fill)
        c.border = _thin_border()
    ws.row_dimensions[row].height = 24


def _data_row(ws, row, values, money_cols=None, bold_first=False, bg=None):
    money_cols = money_cols or []
    for i, v in enumerate(values, start=1):
        c = ws.cell(row, i, v)
        if i == 1:
            c.font = _font(size=10, bold=bold_first)
            c.alignment = _left()
        elif i in money_cols:
            c.font = _font(size=10)
            c.alignment = _right()
            if isinstance(v, (int, float)):
                c.number_format = '#,##0.00;[Red]-#,##0.00'
        else:
            c.font = _font(size=10)
            c.alignment = _left()
        c.border = _thin_border()
        if bg:
            c.fill = _fill(bg)


def _total_row(ws, row, label, values, money_cols, fill=PALE_LAVENDER):
    cells = [label] + list(values)
    for i, v in enumerate(cells, start=1):
        c = ws.cell(row, i, v)
        c.font = _font(size=10, bold=True, color="FF001441")
        if i == 1:
            c.alignment = _left()
        elif i in money_cols:
            c.alignment = _right()
            if isinstance(v, (int, float)):
                c.number_format = '#,##0.00;[Red]-#,##0.00'
        c.fill = _fill(fill)
        c.border = _thin_border()


# ---- Sheet builders ------------------------------------------------------

def _write_summary(wb, firm: Dict, period_label: str, gstr2b: Dict, comp: Dict):
    ws = wb.create_sheet("Summary")
    _set_col_widths(ws, [38, 18, 18, 18, 18])

    _title_row(ws, 1, "GSTR-3B  -  Liability Computation Summary", 5, size=16, height=32)

    r = 3
    ws.cell(r, 1, "Firm").font = _font(bold=True, color="FF6B7280", size=10)
    ws.cell(r, 2, firm.get("name", "")).font = _font(bold=True, size=11)
    r += 1
    ws.cell(r, 1, "GSTIN").font = _font(bold=True, color="FF6B7280", size=10)
    ws.cell(r, 2, firm.get("gstin", "")).font = _font(size=11)
    r += 1
    ws.cell(r, 1, "Return Period").font = _font(bold=True, color="FF6B7280", size=10)
    ws.cell(r, 2, period_label).font = _font(size=11)
    r += 1
    ws.cell(r, 1, "Generated On").font = _font(bold=True, color="FF6B7280", size=10)
    ws.cell(r, 2, datetime.now().strftime("%d-%b-%Y %H:%M")).font = _font(size=11)
    r += 2

    # KPI strip — totals
    _section_row(ws, r, "Period Totals", 5)
    r += 1
    _table_header(ws, r, ["Metric", "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)"])
    r += 1

    rows = [
        ("Output Tax (Table 3.1)",        comp["output_tax"]),
        ("ITC Available (Table 4A)",      comp["itc_available"]),
        ("ITC Reversal (Table 4B)",       comp["itc_reversal"]),
        ("Net ITC for the period",        comp["net_itc"]),
        ("Opening Credit Ledger",         comp["opening_balance"]),
        ("Credit Used (Set-off)",         comp["credit_used"]),
        ("Cash Payable",                  comp["cash_payable"]),
        ("Closing Credit Ledger",         comp["closing_balance"]),
    ]
    for label, t in rows:
        is_cash = label == "Cash Payable"
        is_closing = label == "Closing Credit Ledger"
        bg = "FFFEF3C7" if is_cash else ("FFDCFCE7" if is_closing else None)
        _data_row(ws, r, [label, t["igst"], t["cgst"], t["sgst"], t["cess"]],
                  money_cols={2, 3, 4, 5}, bold_first=is_cash or is_closing, bg=bg)
        r += 1

    # Grand totals
    r += 1
    _section_row(ws, r, "Grand Totals (all heads combined)", 5)
    r += 1
    _table_header(ws, r, ["Description", "Amount (₹)", "", "", ""])
    r += 1
    grand = [
        ("Total Output Tax",         comp["total_output"]),
        ("Total ITC + Opening",      comp["total_itc"]),
        ("Total Credit Used",        comp["total_credit_used"]),
        ("Total Cash Payable",       comp["total_cash_payable"]),
    ]
    for label, amt in grand:
        is_cash = label == "Total Cash Payable"
        bg = "FFFEF3C7" if is_cash else None
        for col in range(1, 6):
            c = ws.cell(r, col)
            c.border = _thin_border()
            if bg:
                c.fill = _fill(bg)
        ws.cell(r, 1, label).font = _font(size=10, bold=True)
        ws.cell(r, 1).alignment = _left()
        ws.cell(r, 2, amt).font = _font(size=11, bold=is_cash, color=RED if is_cash else "FF1A1A2E")
        ws.cell(r, 2).alignment = _right()
        ws.cell(r, 2).number_format = '#,##0.00'
        # merge across remaining cols for cleanliness
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        r += 1

    # Footer
    r += 2
    note = ws.cell(r, 1,
        "This is a computational worksheet. The actual GSTR-3B return must be "
        "filed on the GST portal. Verify all figures before filing.")
    note.font = _font(size=9, italic=True, color=GREY_MUTED)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

    ws.sheet_view.showGridLines = False


def _write_itc_details(wb, gstr2b: Dict):
    ws = wb.create_sheet("ITC Details (GSTR-2B)")
    _set_col_widths(ws, [40, 18, 18, 18, 18])
    _title_row(ws, 1, "ITC Details from GSTR-2B", 5)
    r = 3

    categories = [
        ("all_other_itc",   "All other ITC (B2B — Table 4(A)(5))"),
        ("reverse_charge",  "Inward supplies — Reverse charge (Table 3.1(d)/4(A)(3))"),
        ("isd",             "Inward supplies from ISD (Table 4(A)(4))"),
        ("imports",         "Import of goods (Table 4(A)(1))"),
        ("credit_notes",    "Credit notes (Part B — netting reduction)"),
    ]

    for sheet_key, sheet_label, color in [
        ("itc_available",    "ITC Available", PURPLE),
        ("itc_not_available", "ITC Not Available", "FFB91C1C"),
        ("itc_reversal",     "ITC Reversal", "FFD97706"),
        ("itc_rejected",     "ITC Rejected (IMS)", "FF6B7280"),
    ]:
        _section_row(ws, r, sheet_label, 5, fill=PURPLE_LIGHT)
        r += 1
        _table_header(ws, r, ["Category", "IGST", "CGST", "SGST", "Cess"], fill=color)
        r += 1
        section_data = gstr2b.get(sheet_key, {})
        for cat_key, cat_label in categories:
            tax = section_data.get(cat_key, {"igst": 0, "cgst": 0, "sgst": 0, "cess": 0})
            _data_row(ws, r, [cat_label, tax["igst"], tax["cgst"], tax["sgst"], tax["cess"]],
                      money_cols={2, 3, 4, 5})
            r += 1
        # Total
        total = section_data.get("total", {"igst": 0, "cgst": 0, "sgst": 0, "cess": 0})
        _total_row(ws, r, "Net Total",
                   [total["igst"], total["cgst"], total["sgst"], total["cess"]],
                   money_cols={2, 3, 4, 5})
        r += 2

    ws.sheet_view.showGridLines = False


def _write_output_tax(wb, comp: Dict):
    ws = wb.create_sheet("Output Tax (Table 3.1)")
    _set_col_widths(ws, [40, 22])
    _title_row(ws, 1, "Output Tax — Table 3.1", 2)
    r = 3
    _table_header(ws, r, ["Tax Head", "Amount (₹)"])
    r += 1
    out = comp["output_tax"]
    for k, label in [("igst", "Integrated Tax (IGST)"),
                     ("cgst", "Central Tax (CGST)"),
                     ("sgst", "State/UT Tax (SGST/UTGST)"),
                     ("cess", "Cess")]:
        _data_row(ws, r, [label, out[k]], money_cols={2})
        r += 1
    _total_row(ws, r, "Total Output Tax",
               [comp["total_output"]], money_cols={2})
    ws.sheet_view.showGridLines = False


def _write_itc_computation(wb, comp: Dict):
    ws = wb.create_sheet("ITC Computation (Table 4)")
    _set_col_widths(ws, [40, 18, 18, 18, 18])
    _title_row(ws, 1, "ITC Computation — Table 4", 5)
    r = 3
    _table_header(ws, r, ["Item", "IGST", "CGST", "SGST", "Cess"])
    r += 1

    rows = [
        ("4(A)  ITC Available (gross)",  comp["itc_available"], None),
        ("4(B)  ITC to be Reversed",     comp["itc_reversal"], None),
        ("4(C)  Net ITC for the period", comp["net_itc"], PALE_LAVENDER),
        ("Opening Credit Ledger Balance", comp["opening_balance"], None),
        ("Total Credit Available",        comp["credit_pool"], PURPLE_LIGHT),
    ]
    for label, t, bg in rows:
        is_total = label.startswith("Total Credit")
        _data_row(ws, r, [label, t["igst"], t["cgst"], t["sgst"], t["cess"]],
                  money_cols={2, 3, 4, 5}, bold_first=is_total, bg=bg)
        r += 1
    ws.sheet_view.showGridLines = False


def _write_setoff(wb, comp: Dict):
    ws = wb.create_sheet("Tax Set-off")
    _set_col_widths(ws, [6, 14, 14, 16, 50])
    _title_row(ws, 1, "Tax Set-off Trail  (per Section 49 + Rule 88A)", 5)
    r = 3
    _section_row(ws, r, "Order of credit utilisation", 5)
    r += 1
    note = ws.cell(r, 1,
        "1. IGST credit is fully exhausted before CGST/SGST credit can be used.\n"
        "2. CGST credit cannot be used to offset SGST liability (and vice versa).\n"
        "3. Cess credit can only be used for Cess liability.")
    note.font = _font(size=9, italic=True, color=GREY_MUTED)
    note.alignment = _left()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 50
    r += 2

    _table_header(ws, r, ["#", "From Credit", "To Liability", "Amount (₹)", "Rule / Note"])
    r += 1

    if not comp.get("setoff_steps"):
        ws.cell(r, 1, "No set-off needed (no liability or no credit).").font = _font(italic=True, color=GREY_MUTED)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    else:
        for i, step in enumerate(comp["setoff_steps"], start=1):
            _data_row(ws, r, [i, step["from"], step["to"], step["amount"], step["rule"]],
                      money_cols={4})
            ws.cell(r, 2).alignment = _center()
            ws.cell(r, 3).alignment = _center()
            r += 1
        _total_row(ws, r, "Total Credit Utilised",
                   ["", "", comp["total_credit_used"], ""],
                   money_cols={4})

    ws.sheet_view.showGridLines = False


def _write_cash(wb, comp: Dict):
    ws = wb.create_sheet("Cash Payable")
    _set_col_widths(ws, [40, 22])
    _title_row(ws, 1, "Cash Payable  (after credit set-off)", 2, fill=NIGHT_BLUE)
    r = 3
    _table_header(ws, r, ["Tax Head", "Cash to Pay (₹)"])
    r += 1
    cp = comp["cash_payable"]
    for k, label in [("igst", "IGST"), ("cgst", "CGST"),
                     ("sgst", "SGST/UTGST"), ("cess", "Cess")]:
        amt = cp[k]
        bg = "FFFEF3C7" if amt > 0 else "FFDCFCE7"
        _data_row(ws, r, [label, amt], money_cols={2}, bg=bg)
        r += 1
    _total_row(ws, r, "Total Cash Payable",
               [comp["total_cash_payable"]], money_cols={2}, fill="FFFEF3C7")

    r += 3
    note = ws.cell(r, 1,
        "Pay this amount through the GST portal using PMT-06 / electronic cash ledger.")
    note.font = _font(size=10, italic=True, color=GREY_MUTED)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    ws.sheet_view.showGridLines = False


def _write_ledger(wb, comp: Dict):
    ws = wb.create_sheet("Credit Ledger")
    _set_col_widths(ws, [40, 18, 18, 18, 18])
    _title_row(ws, 1, "Electronic Credit Ledger Movement", 5)
    r = 3
    _table_header(ws, r, ["Movement", "IGST", "CGST", "SGST", "Cess"])
    r += 1

    rows = [
        ("Opening Balance",        comp["opening_balance"], None),
        ("Add: Net ITC this month", comp["net_itc"], None),
        ("Sub-total (Pool)",        comp["credit_pool"], PURPLE_LIGHT),
        ("Less: Used in Set-off",   comp["credit_used"], None),
        ("Closing Balance",         comp["closing_balance"], "FFDCFCE7"),
    ]
    for label, t, bg in rows:
        bold = label in ("Sub-total (Pool)", "Closing Balance")
        _data_row(ws, r, [label, t["igst"], t["cgst"], t["sgst"], t["cess"]],
                  money_cols={2, 3, 4, 5}, bold_first=bold, bg=bg)
        r += 1
    ws.sheet_view.showGridLines = False


# ---- Public --------------------------------------------------------------

def write_gstr3b_excel(out_path: str | Path,
                       firm: Dict[str, Any],
                       period_label: str,
                       gstr2b: Dict[str, Any],
                       computation: Dict[str, Any]) -> Path:
    """
    Build the GSTR-3B workbook and save to `out_path`.

    Returns the Path object of the saved file.
    """
    wb = openpyxl.Workbook()
    # remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    _write_summary(wb, firm, period_label, gstr2b, computation)
    _write_itc_details(wb, gstr2b)
    _write_output_tax(wb, computation)
    _write_itc_computation(wb, computation)
    _write_setoff(wb, computation)
    _write_cash(wb, computation)
    _write_ledger(wb, computation)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
