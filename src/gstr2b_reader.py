"""
GSTR-2B reader — parses the Excel file the GST portal exports.

The summary sheets (ITC Available / Not Available / Reversal / Rejected) all
share the same column layout:
    A: S.no.                B: Heading              C: GSTR-3B table
    D: Integrated Tax       E: Central Tax          F: State/UT Tax
    G: Cess                 H: Advisory

Heading rows with a Roman numeral (I, II, III, IV) in column A are the
*category roll-ups* — these are the numbers we use for GSTR-3B Table 4 entries.

Public API:
    parse_gstr2b(path)  ->  dict (see _empty_result)
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


# --- Helpers ---------------------------------------------------------------

def _safe_float(v: Any) -> float:
    """Cell value -> float, never raises. Blank cells become 0."""
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _empty_tax() -> Dict[str, float]:
    return {"igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}


def _row_to_tax(ws, row: int) -> Dict[str, float]:
    """Extract IGST/CGST/SGST/Cess values from a single row (cols D-G)."""
    return {
        "igst": _safe_float(ws.cell(row, 4).value),
        "cgst": _safe_float(ws.cell(row, 5).value),
        "sgst": _safe_float(ws.cell(row, 6).value),
        "cess": _safe_float(ws.cell(row, 7).value),
    }


def _add_tax(a: Dict[str, float], b: Dict[str, float]) -> Dict[str, float]:
    return {k: round(a[k] + b[k], 2) for k in a}


def _sub_tax(a: Dict[str, float], b: Dict[str, float]) -> Dict[str, float]:
    return {k: round(a[k] - b[k], 2) for k in a}


# --- Category detection ----------------------------------------------------

# Map S.No (column A) values to internal category keys. Roman numerals are
# what the portal uses. We only care about Part A rows for "ITC to claim";
# Part B rows are credit notes that net off.
CATEGORY_LABELS = {
    "I":   "all_other_itc",          # Table 4(A)(5) - regular B2B
    "II":  "isd",                    # Table 4(A)(4) - ISD
    "III": "reverse_charge",         # Table 3.1(d) + 4(A)(3)
    "IV":  "imports",                # Table 4(A)(1) - IMPG + IMPGSEZ
}


def _parse_summary_sheet(ws) -> Dict[str, Dict[str, float]]:
    """
    Read a summary-style sheet and return a dict keyed by category.

    Returns: {
        "all_other_itc":  {"igst": .., "cgst": .., "sgst": .., "cess": ..},
        "isd":            {...},
        "reverse_charge": {...},
        "imports":        {...},
        "credit_notes":   {...},   # Part B total (negative effect on ITC)
        "total":          {...},   # whole-sheet total
    }
    """
    result: Dict[str, Dict[str, float]] = {
        k: _empty_tax() for k in CATEGORY_LABELS.values()
    }
    result["credit_notes"] = _empty_tax()
    result["total"] = _empty_tax()

    in_part_b = False

    for row in range(7, ws.max_row + 1):
        a = ws.cell(row, 1).value
        if a is None:
            continue
        a_str = str(a).strip()

        # Detect Part B (credit notes section)
        if a_str.startswith("Part B"):
            in_part_b = True
            continue
        if a_str.startswith("Part A"):
            in_part_b = False
            continue

        if a_str in CATEGORY_LABELS:
            tax = _row_to_tax(ws, row)
            if in_part_b:
                # All Part B roll-ups go into "credit_notes"
                result["credit_notes"] = _add_tax(result["credit_notes"], tax)
            else:
                key = CATEGORY_LABELS[a_str]
                result[key] = tax

    # Compute net total (Part A categories minus credit notes)
    parta_total = _empty_tax()
    for k in CATEGORY_LABELS.values():
        parta_total = _add_tax(parta_total, result[k])
    result["total"] = _sub_tax(parta_total, result["credit_notes"])

    return result


# --- Public ---------------------------------------------------------------

def _empty_result() -> Dict[str, Any]:
    return {
        "filename": "",
        "gstin": "",
        "period": "",
        "generated_on": "",
        "itc_available": {
            "all_other_itc": _empty_tax(),
            "isd": _empty_tax(),
            "reverse_charge": _empty_tax(),
            "imports": _empty_tax(),
            "credit_notes": _empty_tax(),
            "total": _empty_tax(),
        },
        "itc_not_available": {
            "all_other_itc": _empty_tax(),
            "isd": _empty_tax(),
            "reverse_charge": _empty_tax(),
            "imports": _empty_tax(),
            "credit_notes": _empty_tax(),
            "total": _empty_tax(),
        },
        "itc_reversal": {
            "all_other_itc": _empty_tax(),
            "isd": _empty_tax(),
            "reverse_charge": _empty_tax(),
            "imports": _empty_tax(),
            "credit_notes": _empty_tax(),
            "total": _empty_tax(),
        },
        "itc_rejected": {
            "all_other_itc": _empty_tax(),
            "isd": _empty_tax(),
            "reverse_charge": _empty_tax(),
            "imports": _empty_tax(),
            "credit_notes": _empty_tax(),
            "total": _empty_tax(),
        },
        "table4": {},
        "invoices": [],  # Per-invoice detail rows (Phase 4a)
        "warnings": [],
    }


def _extract_metadata(ws_readme) -> Dict[str, str]:
    """Pull GSTIN / period / generated date from the Read me sheet."""
    meta = {"gstin": "", "period": "", "generated_on": ""}
    for r in range(1, min(50, ws_readme.max_row + 1)):
        for c in range(1, min(8, ws_readme.max_column + 1)):
            v = ws_readme.cell(r, c).value
            if not v:
                continue
            s = str(v).strip()
            if "GSTIN" in s and ":" in s:
                meta["gstin"] = s.split(":", 1)[1].strip()
            elif "Return period" in s.lower() or "tax period" in s.lower():
                # Next cell or part after colon
                if ":" in s:
                    meta["period"] = s.split(":", 1)[1].strip()
            elif "generated" in s.lower() and ":" in s.lower():
                if ":" in s:
                    meta["generated_on"] = s.split(":", 1)[1].strip()
    return meta


def parse_gstr2b(path: str | Path) -> Dict[str, Any]:
    """
    Parse a GSTR-2B Excel file from the GST portal.

    Returns a dict with parsed totals — see _empty_result() for shape.
    """
    path = Path(path)
    result = _empty_result()
    result["filename"] = path.name

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception as e:
        result["warnings"].append(f"Could not open Excel file: {e}")
        return result

    # Optional metadata
    if "Read me" in wb.sheetnames:
        try:
            meta = _extract_metadata(wb["Read me"])
            result.update(meta)
        except Exception:
            pass

    # Parse the four summary sheets
    sheet_map = {
        "ITC Available":     "itc_available",
        "ITC not available": "itc_not_available",
        "ITC Reversal":      "itc_reversal",
        "ITC Rejected":      "itc_rejected",
    }

    for sheet_name, result_key in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            result["warnings"].append(f"Sheet '{sheet_name}' missing from file")
            continue
        try:
            result[result_key] = _parse_summary_sheet(wb[sheet_name])
        except Exception as e:
            result["warnings"].append(f"Could not parse '{sheet_name}': {e}")

    # Phase 4a: extract per-invoice detail rows from B2B/B2BA/CDNR/CDNRA/IMPG/IMPGSEZ
    try:
        result["invoices"] = _extract_invoice_details(wb)
    except Exception as e:
        result["warnings"].append(f"Could not extract invoice list: {e}")

    # Build the formal GSTR-3B Table 4 line breakdown.
    # Reference: Notification 14/2022-CT (Oct 2022) — restructured Table 4.
    try:
        result["table4"] = _build_table4_lines(result)
    except Exception as e:
        result["warnings"].append(f"Could not build Table 4 breakdown: {e}")

    # Section 16(4): ITC on any prior-FY invoice cannot be claimed after
    # 30 November of the following FY (or the date of filing of the annual
    # return, whichever is earlier). We scan detail sheets for invoice dates
    # belonging to a prior FY relative to the 2B period and flag them.
    try:
        prior_fy_count, examples = _scan_prior_fy_invoices(wb, result.get("period", ""))
        if prior_fy_count:
            result["warnings"].append(
                f"Section 16(4): {prior_fy_count} invoice(s) found from a prior "
                f"financial year. ITC cannot be claimed after 30-Nov of the FY "
                f"following the invoice's FY. Examples: {', '.join(examples[:3])}"
            )
    except Exception as e:
        result["warnings"].append(f"Could not run Section 16(4) check: {e}")

    return result


# --- Per-invoice detail extraction (Phase 4a) -----------------------------

# Map detail-sheet name -> internal category. These names match the GSTN portal
# Excel sheet titles exactly (case-sensitive after .strip().upper()).
_DETAIL_SHEET_CATEGORY = {
    "B2B":     "all_other_itc",
    "B2BA":    "all_other_itc",   # amendments
    "CDNR":    "credit_notes",
    "CDNRA":   "credit_notes",    # amendments
    "IMPG":    "imports",
    "IMPGSEZ": "imports",
    "ISD":     "isd",
    "ISDA":    "isd",
}

# Synonyms for header lookups (case-insensitive substring match).
# Keyword lists for fuzzy header matching. Order matters within each tuple —
# more specific phrases first so e.g. "state/ut tax" beats a bare "tax".
_HDR = {
    "supplier_gstin":   ("gstin of supplier", "gstin/uin of supplier", "supplier gstin"),
    "supplier_name":    ("trade/legal name", "trade name of the supplier",
                         "legal name of the supplier", "supplier name"),
    "invoice_no":       ("invoice number", "note number", "doc number",
                         "shipping bill number"),
    "invoice_date":     ("invoice date", "note date", "doc date", "shipping bill date"),
    "invoice_type":     ("invoice type", "note type", "doc type", "supply type"),
    "invoice_value":    ("invoice value", "note value", "doc value"),
    "place_of_supply":  ("place of supply",),
    "reverse_charge":   ("supply attract reverse charge", "reverse charge"),
    "rate":             ("rate(%)", "rate (%)", "applicable % of tax rate"),
    "taxable_value":    ("taxable value",),
    "igst":             ("integrated tax", "igst"),
    "cgst":             ("central tax", "cgst"),
    "sgst":             ("state/ut tax", "state ut tax", "state tax", "sgst"),
    "cess":             ("cess(", "cess (", " cess"),  # avoid bare 'cess' substring
    "itc_availability": ("itc availability",),
    "reason":           ("reason", "remarks"),
    "source":           ("source",),  # IMPG: port of import etc.
}


def _extract_invoice_details(wb) -> List[Dict[str, Any]]:
    """
    Walk every detail sheet present in the workbook and emit a flat list
    of per-invoice dicts. Each row is uniquely identified by
    (category, sheet, source_row).
    """
    invoices: List[Dict[str, Any]] = []
    seq = 0
    for sheet_name in wb.sheetnames:
        key = sheet_name.strip().upper()
        category = _DETAIL_SHEET_CATEGORY.get(key)
        if not category:
            continue
        ws = wb[sheet_name]
        header_row = _find_header_row(ws)
        if header_row is None:
            continue
        # GSTR-2B detail sheets have merged multi-row headers. Build a column
        # map by concatenating up to 3 header rows so "Integrated Tax (₹)"
        # split across rows still matches.
        col_map = {field: _find_col_band(ws, header_row, kws)
                   for field, kws in _HDR.items()}
        # Find the actual first data row (skip blank/unit/sub-header rows just
        # below the title row).
        data_start = _find_data_start(ws, header_row, col_map)
        for r in range(data_start, ws.max_row + 1):
            # Skip blank rows — require at least supplier GSTIN or invoice number
            sup = ws.cell(r, col_map["supplier_gstin"]).value if col_map["supplier_gstin"] else None
            inum = ws.cell(r, col_map["invoice_no"]).value if col_map["invoice_no"] else None
            if not sup and not inum:
                continue
            seq += 1
            inv = {
                "id": f"{key}-{seq}",
                "category": category,
                "source_sheet": key,
                "supplier_gstin":   _cell_str(ws, r, col_map["supplier_gstin"]),
                "supplier_name":    _cell_str(ws, r, col_map["supplier_name"]),
                "invoice_no":       _cell_str(ws, r, col_map["invoice_no"]),
                "invoice_date":     _cell_date_str(ws, r, col_map["invoice_date"]),
                "invoice_type":     _cell_str(ws, r, col_map["invoice_type"]),
                "invoice_value":    _safe_float(_cell_raw(ws, r, col_map["invoice_value"])),
                "place_of_supply":  _cell_str(ws, r, col_map["place_of_supply"]),
                "reverse_charge":   _cell_str(ws, r, col_map["reverse_charge"]),
                "rate":             _safe_float(_cell_raw(ws, r, col_map["rate"])),
                "taxable_value":    _safe_float(_cell_raw(ws, r, col_map["taxable_value"])),
                "igst":             _safe_float(_cell_raw(ws, r, col_map["igst"])),
                "cgst":             _safe_float(_cell_raw(ws, r, col_map["cgst"])),
                "sgst":             _safe_float(_cell_raw(ws, r, col_map["sgst"])),
                "cess":             _safe_float(_cell_raw(ws, r, col_map["cess"])),
                "itc_availability": _cell_str(ws, r, col_map["itc_availability"]),
                "reason":           _cell_str(ws, r, col_map["reason"]),
            }
            # Credit notes carry negative sign in portal — preserve as-is.
            invoices.append(inv)
    return invoices


def _cell_raw(ws, row: int, col: Optional[int]):
    if col is None:
        return None
    return ws.cell(row, col).value


def _cell_str(ws, row: int, col: Optional[int]) -> str:
    v = _cell_raw(ws, row, col)
    if v is None:
        return ""
    return str(v).strip()


def _cell_date_str(ws, row: int, col: Optional[int]) -> str:
    v = _cell_raw(ws, row, col)
    d = _to_date(v)
    if d:
        return d.strftime("%d-%m-%Y")
    return _cell_str(ws, row, col)


# --- Table 4 line breakdown (post-Oct 2022) -------------------------------

def _build_table4_lines(parsed: Dict[str, Any]) -> Dict[str, Dict[str, float]]:
    """
    Produce the official GSTR-3B Table 4 line items from parsed GSTR-2B.

      4(A) ITC Available:
        4(A)(1) Import of goods            <- itc_available.imports
        4(A)(2) Import of services         <- (manual; not in 2B)
        4(A)(3) Inward supplies (RCM)      <- itc_available.reverse_charge
        4(A)(4) Inward supplies from ISD   <- itc_available.isd
        4(A)(5) All other ITC              <- itc_available.all_other_itc
                                              minus credit_notes (Part B)

      4(B) ITC Reversed:
        4(B)(1) As per Rules 38,42,43, Sec 17(5)  <- (manual)
        4(B)(2) Others (temporary reversals)      <- itc_reversal.total

      4(C) Net ITC available = 4(A) - 4(B)

      4(D) Other details:
        4(D)(1) ITC reclaimed (from 4(B)(2))      <- (manual)
        4(D)(2) Ineligible ITC u/s 16(4) & POS    <- itc_not_available.total
                                                     + itc_rejected.total
    """
    avail = parsed.get("itc_available", {})
    notavail = parsed.get("itc_not_available", {})
    reversal = parsed.get("itc_reversal", {})
    rejected = parsed.get("itc_rejected", {})

    # 4(A)(5) = all_other_itc net of Part B credit notes
    a5_net = _sub_tax(
        avail.get("all_other_itc", _empty_tax()),
        avail.get("credit_notes", _empty_tax()),
    )

    table4 = {
        "4A1_import_goods":        dict(avail.get("imports", _empty_tax())),
        "4A2_import_services":     _empty_tax(),  # manual
        "4A3_reverse_charge":      dict(avail.get("reverse_charge", _empty_tax())),
        "4A4_isd":                 dict(avail.get("isd", _empty_tax())),
        "4A5_all_other_itc":       a5_net,
        "4B1_rules_38_42_43_17_5": _empty_tax(),  # manual
        "4B2_others":              dict(reversal.get("total", _empty_tax())),
        "4D1_reclaimed":           _empty_tax(),  # manual
        "4D2_ineligible_16_4_pos": _add_tax(
            notavail.get("total", _empty_tax()),
            rejected.get("total", _empty_tax()),
        ),
    }

    # 4(A) total and 4(C) net
    a_total = _empty_tax()
    for k in ("4A1_import_goods", "4A2_import_services", "4A3_reverse_charge",
              "4A4_isd", "4A5_all_other_itc"):
        a_total = _add_tax(a_total, table4[k])
    b_total = _add_tax(table4["4B1_rules_38_42_43_17_5"], table4["4B2_others"])
    table4["4A_total"] = a_total
    table4["4B_total"] = b_total
    table4["4C_net_itc"] = _sub_tax(a_total, b_total)
    return table4


# --- Section 16(4) helpers ------------------------------------------------

def _fy_of(d: date) -> int:
    """Indian financial year start year for a date (Apr-Mar)."""
    return d.year if d.month >= 4 else d.year - 1


def _period_fy(period: str) -> Optional[int]:
    """GSTR-2B period is 'MMYYYY' (e.g. '052026'). Return FY start year."""
    s = (period or "").strip()
    if len(s) == 6 and s.isdigit():
        mm, yyyy = int(s[:2]), int(s[2:])
        return yyyy if mm >= 4 else yyyy - 1
    return None


def _scan_prior_fy_invoices(wb, period: str) -> tuple[int, list[str]]:
    """
    Scan B2B / B2BA / CDNR / IMPG detail sheets for invoice dates that fall in
    an FY earlier than the 2B period's FY. Returns (count, sample_strings).
    """
    period_fy = _period_fy(period)
    if period_fy is None:
        return 0, []

    DETAIL_SHEETS = ("B2B", "B2BA", "CDNR", "CDNRA", "IMPG", "IMPGSEZ")
    count = 0
    samples: list[str] = []
    for name in wb.sheetnames:
        if name.strip().upper() not in DETAIL_SHEETS:
            continue
        ws = wb[name]
        # Find the date column by header text (typically "Invoice date" or "Note date")
        header_row = _find_header_row(ws)
        if header_row is None:
            continue
        date_col = _find_col(ws, header_row, ("invoice date", "note date", "doc date"))
        inum_col = _find_col(ws, header_row, ("invoice number", "note number", "doc number"))
        if date_col is None:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(r, date_col).value
            d = _to_date(v)
            if d is None:
                continue
            if _fy_of(d) < period_fy:
                count += 1
                if len(samples) < 5 and inum_col:
                    inum = ws.cell(r, inum_col).value
                    samples.append(f"{inum} dt {d.strftime('%d-%m-%Y')}")
    return count, samples


def _find_header_row(ws) -> Optional[int]:
    """
    Detail sheets have a multi-row merged header. We want the TOP header row
    (containing the merged group titles like 'Integrated Tax'). Heuristic:
    scan first ~15 rows; pick the first row that contains 'GSTIN' anywhere.
    """
    for r in range(1, min(20, ws.max_row + 1)):
        for c in range(1, min(30, ws.max_column + 1)):
            v = ws.cell(r, c).value
            if v and "GSTIN" in str(v).upper():
                return r
    return None


def _find_col(ws, header_row: int, keywords: tuple[str, ...]) -> Optional[int]:
    """Single-row column finder (used by Sec 16(4) scan)."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if not v:
            continue
        s = str(v).lower()
        if any(k in s for k in keywords):
            return c
    return None


def _find_col_band(ws, header_row: int, keywords: tuple[str, ...],
                   band: int = 3) -> Optional[int]:
    """
    Multi-row column finder for merged headers. Concatenates cell text from
    header_row .. header_row + band-1 for each column, then matches keywords.
    Handles merged-cell ranges by replicating the merged value across columns.
    """
    max_col = ws.max_column or 1
    # Build a per-column composite header string by reading `band` rows and
    # honoring merged-cell groups (only top-left has the value in openpyxl).
    composite: Dict[int, str] = {c: "" for c in range(1, max_col + 1)}
    # First, expand merged ranges so every covered cell carries its top-left value
    merged_values: Dict[tuple, str] = {}
    for mr in ws.merged_cells.ranges:
        top_left = ws.cell(mr.min_row, mr.min_col).value
        if top_left is None:
            continue
        s = str(top_left)
        for rr in range(mr.min_row, mr.max_row + 1):
            for cc in range(mr.min_col, mr.max_col + 1):
                merged_values[(rr, cc)] = s
    for dr in range(band):
        r = header_row + dr
        if r > ws.max_row:
            break
        for c in range(1, max_col + 1):
            v = merged_values.get((r, c))
            if v is None:
                v = ws.cell(r, c).value
            if v is None:
                continue
            composite[c] += " " + str(v)
    for c, txt in composite.items():
        s = txt.lower()
        if any(k in s for k in keywords):
            return c
    return None


def _find_data_start(ws, header_row: int, col_map: Dict[str, Optional[int]]) -> int:
    """
    Find the first row below the header band that contains real data — i.e.
    a row where the supplier-GSTIN or invoice-no column has a value that
    LOOKS like data (not another header word).
    """
    sup_col = col_map.get("supplier_gstin")
    inum_col = col_map.get("invoice_no")
    for r in range(header_row + 1, min(header_row + 8, ws.max_row + 1)):
        sup = ws.cell(r, sup_col).value if sup_col else None
        inum = ws.cell(r, inum_col).value if inum_col else None
        for cand in (sup, inum):
            if cand is None:
                continue
            s = str(cand).strip()
            if not s:
                continue
            # Skip rows that look like sub-headers (contain header keywords)
            low = s.lower()
            if any(w in low for w in ("gstin", "invoice", "trade", "legal",
                                      "place of supply", "(₹)", "(rs)")):
                continue
            return r
    return header_row + 1


def _to_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None
