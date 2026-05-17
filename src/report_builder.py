"""
Excel report builder.
Generates a multi-sheet workbook with:
  - Summary
  - Invoice-wise list (consolidated)
  - HSN-wise summary
  - B2B / B2CL / B2CS classification
  - Exceptions / GSTIN corrections
  - Document Issued summary
"""
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F4E78")
NORMAL_FONT = Font(name="Arial", size=10)
ISSUE_FILL = PatternFill("solid", start_color="FFEB99")
THIN = Side(border_style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row_num, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _autofit(ws, max_width=50):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), max_width))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def _write_table(ws, start_row, headers, rows):
    for i, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=i, value=h)
    _style_header(ws, start_row, len(headers))
    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = NORMAL_FONT
            cell.border = BORDER
    return start_row + len(rows) + 1


def build_report(
    firm_name: str,
    firm_gstin: str,
    return_period: str,
    invoices: list,
    buckets: dict,
    exceptions_df: pd.DataFrame,
    output_path: str,
    excluded_invoices: list = None,
):
    wb = Workbook()

    # ========== Summary ==========
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"GSTR-1 Summary — {firm_name}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    summary_rows = [
        ("Firm GSTIN", firm_gstin),
        ("Return Period (MMYYYY)", return_period),
        ("Total Invoices", len(invoices)),
        ("B2B Invoices", len(buckets.get("b2b", []))),
        ("B2CL Invoices", len(buckets.get("b2cl", []))),
        ("B2CS Invoices", len(buckets.get("b2cs", []))),
        ("Total Taxable Value", round(sum(i["invoice_total_taxable"] for i in invoices), 2)),
        ("Total Tax", round(sum(i["invoice_total_tax"] for i in invoices), 2)),
        ("Total Invoice Value", round(sum(i["invoice_value"] for i in invoices), 2)),
        ("Exceptions Flagged", len(exceptions_df) if exceptions_df is not None else 0),
    ]
    for i, (k, v) in enumerate(summary_rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(name="Arial", bold=True)
        ws.cell(row=i, column=2, value=v).font = NORMAL_FONT
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25

    # ========== Invoice-wise summary ==========
    ws = wb.create_sheet("Invoice Summary")
    headers = [
        "Invoice No", "Date", "GSTIN", "Customer", "Type", "POS",
        "Items", "Taxable Value", "IGST", "CGST", "SGST", "Cess",
        "Invoice Value",
    ]
    rows = []
    for inv in invoices:
        if inv["is_b2b"]:
            inv_type = "B2B"
        elif inv["is_interstate"] and inv["invoice_value"] > 250000:
            inv_type = "B2CL"
        else:
            inv_type = "B2CS"
        igst = sum(i["igst_amount"] for i in inv["items"])
        cgst = sum(i["cgst_amount"] for i in inv["items"])
        sgst = sum(i["sgst_amount"] for i in inv["items"])
        cess = sum(i["cess_amount"] for i in inv["items"])
        rows.append([
            inv["invoice_no"],
            inv["invoice_date"].strftime("%d-%m-%Y") if inv["invoice_date"] else "",
            inv["gstin"],
            inv["customer_name"],
            inv_type,
            inv["place_of_supply"],
            len(inv["items"]),
            round(inv["invoice_total_taxable"], 2),
            round(igst, 2), round(cgst, 2), round(sgst, 2), round(cess, 2),
            round(inv["invoice_value"], 2),
        ])
    _write_table(ws, 1, headers, rows)
    _autofit(ws)

    # ========== HSN-wise summary — split into B2B and B2C ==========
    # Matches the official GST Offline Tool V2.2 template format:
    #   hsn(b2b)  → for sales to registered customers
    #   hsn(b2c)  → for sales to unregistered customers (B2CL + B2CS)
    # Columns match the official CSV exactly:
    #   HSN, Description, UQC, Total Quantity, Total Value, Taxable Value,
    #   Integrated Tax Amount, Central Tax Amount, State/UT Tax Amount, Cess Amount, Rate
    from collections import defaultdict

    # Common UQC code → label mapping (from V2.2 master sheet)
    UQC_LABELS = {
        "BAG": "BAG-BAGS", "BAL": "BAL-BALE", "BDL": "BDL-BUNDLES",
        "BKL": "BKL-BUCKLES", "BOU": "BOU-BILLIONS OF UNITS",
        "BOX": "BOX-BOX", "BTL": "BTL-BOTTLES", "BUN": "BUN-BUNCHES",
        "CAN": "CAN-CANS", "CBM": "CBM-CUBIC METERS",
        "CCM": "CCM-CUBIC CENTIMETERS", "CMS": "CMS-CENTIMETERS",
        "CTN": "CTN-CARTONS", "DOZ": "DOZ-DOZENS", "DRM": "DRM-DRUMS",
        "GGK": "GGK-GREAT GROSS", "GMS": "GMS-GRAMMES",
        "GRS": "GRS-GROSS", "GYD": "GYD-GROSS YARDS",
        "KGS": "KGS-KILOGRAMS", "KLR": "KLR-KILOLITRE",
        "KME": "KME-KILOMETRE", "LTR": "LTR-LITRES", "MLT": "MLT-MILILITRE",
        "MTR": "MTR-METERS", "MTS": "MTS-METRIC TON", "NOS": "NOS-NUMBERS",
        "PAC": "PAC-PACKS", "PCS": "PCS-PIECES", "PRS": "PRS-PAIRS",
        "QTL": "QTL-QUINTAL", "ROL": "ROL-ROLLS", "SET": "SET-SETS",
        "SQF": "SQF-SQUARE FEET", "SQM": "SQM-SQUARE METERS",
        "SQY": "SQY-SQUARE YARDS", "TBS": "TBS-TABLETS",
        "TGM": "TGM-TEN GROSS", "THD": "THD-THOUSANDS",
        "TON": "TON-TONNES", "TUB": "TUB-TUBES", "UGS": "UGS-US GALLONS",
        "UNT": "UNT-UNITS", "YDS": "YDS-YARDS", "OTH": "OTH-OTHERS",
        "NA": "NA",
    }

    def _normalize_uqc_for_hsn(uqc: str) -> str:
        """Map a UQC code to its full GSTN label (e.g. 'NOS' → 'NOS-NUMBERS').
        Services (NA) stay as 'NA'."""
        s = str(uqc or "").upper().strip().rstrip(".")
        # Try full label match first
        if s in UQC_LABELS:
            return UQC_LABELS[s]
        # Common variants
        if s in ("NO", "NOS.", "NO.", "NUMBER", "NUMBERS"):
            return UQC_LABELS["NOS"]
        if s in ("KG", "KILOGRAM", "KILOGRAMS"):
            return UQC_LABELS["KGS"]
        if s in ("PC", "PIECE", "PIECES"):
            return UQC_LABELS["PCS"]
        return UQC_LABELS["OTH"]

    def _build_hsn_rows(invoices_subset: list) -> list:
        """Aggregate invoices into HSN-rate rows in official-template format."""
        groups = defaultdict(lambda: {
            "qty": 0.0, "txval": 0.0,
            "iamt": 0.0, "camt": 0.0, "samt": 0.0, "csamt": 0.0,
            "uqc_raw": "",
        })
        for inv in invoices_subset:
            for item in inv["items"]:
                hsn = str(item.get("hsn", "")).strip()
                rate = round(float(item.get("tax_rate", 0)), 2)
                key = (hsn, rate)
                g = groups[key]
                # Excel sheets keep actual UQC and qty for ALL HSN codes
                # (unlike the JSON which uses NA/0 for services).
                g["qty"] += float(item.get("quantity", 0) or 0)
                if not g["uqc_raw"]:
                    g["uqc_raw"] = item.get("uqc", "OTH")
                g["txval"] += float(item.get("taxable_value", 0) or 0)
                g["iamt"] += float(item.get("igst_amount", 0) or 0)
                g["camt"] += float(item.get("cgst_amount", 0) or 0)
                g["samt"] += float(item.get("sgst_amount", 0) or 0)
                g["csamt"] += float(item.get("cess_amount", 0) or 0)

        rows = []
        for (hsn, rate), g in sorted(groups.items()):
            total = g["txval"] + g["iamt"] + g["camt"] + g["samt"] + g["csamt"]
            uqc_label = _normalize_uqc_for_hsn(g["uqc_raw"])
            rows.append([
                hsn,                                      # HSN
                "",                                       # Description (blank, GSTN auto-fills)
                uqc_label,                                # UQC e.g. "NOS-NUMBERS"
                round(g["qty"], 2),                       # Total Quantity (actual, not zeroed for services)
                round(total, 2),                          # Total Value
                round(g["txval"], 2),                     # Taxable Value
                round(g["iamt"], 2),                      # Integrated Tax Amount
                round(g["camt"], 2),                      # Central Tax Amount
                round(g["samt"], 2),                      # State/UT Tax Amount
                round(g["csamt"], 2),                     # Cess Amount
                rate,                                     # Rate
            ])
        return rows

    HSN_HEADERS = [
        "HSN", "Description", "UQC", "Total Quantity", "Total Value",
        "Taxable Value", "Integrated Tax Amount", "Central Tax Amount",
        "State/UT Tax Amount", "Cess Amount", "Rate",
    ]

    # --- HSN B2B ---
    b2b_invoices = buckets.get("b2b", [])
    if b2b_invoices:
        ws = wb.create_sheet("HSN B2B")
        b2b_rows = _build_hsn_rows(b2b_invoices)
        _write_table(ws, 1, HSN_HEADERS, b2b_rows)
        _autofit(ws)

    # --- HSN B2C (B2CL + B2CS combined) ---
    b2c_invoices = buckets.get("b2cl", []) + buckets.get("b2cs", [])
    if b2c_invoices:
        ws = wb.create_sheet("HSN B2C")
        b2c_rows = _build_hsn_rows(b2c_invoices)
        _write_table(ws, 1, HSN_HEADERS, b2c_rows)
        _autofit(ws)

    # If neither B2B nor B2C yielded HSN rows (edge case: empty filing),
    # still create an empty HSN B2B sheet so the report structure is consistent.
    if not b2b_invoices and not b2c_invoices:
        ws = wb.create_sheet("HSN B2B")
        _write_table(ws, 1, HSN_HEADERS, [])
        _autofit(ws)

    # ========== Document Issued ==========
    ws = wb.create_sheet("Document Issued")
    inv_nums = sorted([i["invoice_no"] for i in invoices if i["invoice_no"]])
    rows = []
    if inv_nums:
        rows.append([
            "Invoices for outward supply",
            inv_nums[0],
            inv_nums[-1],
            len(inv_nums),
            0,
            len(inv_nums),
        ])
    _write_table(ws, 1, ["Document Type", "From", "To", "Total", "Cancelled", "Net Issued"], rows)
    _autofit(ws)

    # ========== Exceptions ==========
    ws = wb.create_sheet("Exceptions")
    if exceptions_df is not None and not exceptions_df.empty:
        headers = list(exceptions_df.columns)
        rows = exceptions_df.values.tolist()
        _write_table(ws, 1, headers, rows)
        # Highlight cells with issues
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = ISSUE_FILL
    else:
        ws["A1"] = "No exceptions found."
        ws["A1"].font = NORMAL_FONT
    _autofit(ws)

    # ========== B2B / B2CL / B2CS detail sheets ==========
    for bucket_name in ["b2b", "b2cl", "b2cs"]:
        bucket = buckets.get(bucket_name, [])
        if not bucket:
            continue
        ws = wb.create_sheet(bucket_name.upper())
        headers = [
            "Invoice No", "Date", "GSTIN", "Customer", "POS", "HSN",
            "Description", "UQC", "Qty", "Rate %",
            "Taxable Value", "IGST", "CGST", "SGST", "Cess", "Line Total",
        ]
        rows = []
        for inv in bucket:
            for item in inv["items"]:
                line_total = (item["taxable_value"] + item["igst_amount"]
                              + item["cgst_amount"] + item["sgst_amount"]
                              + item["cess_amount"])
                rows.append([
                    inv["invoice_no"],
                    inv["invoice_date"].strftime("%d-%m-%Y") if inv["invoice_date"] else "",
                    inv["gstin"],
                    inv["customer_name"],
                    inv["place_of_supply"],
                    item["hsn"],
                    item["description"],
                    item["uqc"],
                    item["quantity"],
                    item["tax_rate"],
                    item["taxable_value"],
                    item["igst_amount"],
                    item["cgst_amount"],
                    item["sgst_amount"],
                    item["cess_amount"],
                    round(line_total, 2),
                ])
        _write_table(ws, 1, headers, rows)
        _autofit(ws)

    # ========== Excluded Invoices (Phase 1.2) ==========
    # If invoices were excluded during this generation, dump them into a
    # dedicated sheet so the user can include them in next month's filing.
    if excluded_invoices:
        ws = wb.create_sheet("Excluded Invoices")
        ws["A1"] = f"Invoices Excluded From This Filing — {firm_name}"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:N1")

        ws["A3"] = (
            f"These {len(excluded_invoices)} invoices were excluded by the user "
            f"during JSON generation for return period {return_period}. "
            "Include them in the next month's source data if applicable."
        )
        ws["A3"].font = Font(name="Arial", italic=True, color="808080")
        ws.merge_cells("A3:N3")

        headers = [
            "Excluded On", "Doc Type", "Invoice/Note No", "Date", "GSTIN", "Customer",
            "POS", "Bucket", "HSN/SAC", "UQC", "Quantity", "Rate",
            "Taxable Value", "IGST", "CGST", "SGST", "Cess", "Total Value",
            "Carry-over Status",
        ]
        rows = []
        excluded_on = datetime.now().strftime("%Y-%m-%d %H:%M")
        for inv in excluded_invoices:
            doc_type_label = ("Credit Note" if inv.get("doc_type") == "C"
                              else "Debit Note" if inv.get("doc_type") == "D"
                              else "Invoice")
            bucket = ("B2B" if inv.get("is_b2b") and inv.get("gstin")
                      else "B2CL" if inv.get("is_interstate")
                      else "B2CS")
            inv_date_str = (inv["invoice_date"].strftime("%d-%m-%Y")
                            if inv.get("invoice_date") else "")
            # One row per line item so user can see exactly what to copy back
            for item in inv.get("items", []):
                rows.append([
                    excluded_on,
                    doc_type_label,
                    inv.get("invoice_no", ""),
                    inv_date_str,
                    inv.get("gstin", ""),
                    inv.get("customer_name", ""),
                    inv.get("place_of_supply", ""),
                    bucket,
                    item.get("hsn", ""),
                    item.get("uqc", ""),
                    item.get("quantity", 0),
                    item.get("tax_rate", 0),
                    item.get("taxable_value", 0),
                    item.get("igst_amount", 0),
                    item.get("cgst_amount", 0),
                    item.get("sgst_amount", 0),
                    item.get("cess_amount", 0),
                    round(item.get("taxable_value", 0)
                          + item.get("igst_amount", 0)
                          + item.get("cgst_amount", 0)
                          + item.get("sgst_amount", 0)
                          + item.get("cess_amount", 0), 2),
                    "Pending",  # Phase 2 will track if processed in a later month
                ])
        _write_table(ws, 5, headers, rows)
        _autofit(ws)

    # ====================================================================
    # OFFICIAL TEMPLATE FORMAT SHEETS
    # These sheets exactly match the column structure of the GST Offline
    # Tool V2.2 Excel template, so you can copy-paste them directly into
    # the official offline tool's Excel template if needed.
    # Sheet names match the official template tab names.
    # ====================================================================
    _write_official_template_sheets(wb, firm_gstin, return_period,
                                     invoices, buckets)

    wb.save(output_path)
    return output_path


# ====================================================================
# Official template sheets (V2.2)
# ====================================================================
def _write_official_template_sheets(wb, firm_gstin, return_period,
                                     invoices, buckets):
    """Add sheets in the exact format of the official GSTR-1 Excel template
    (V2.2). Sheet names match the offline tool's tab names verbatim."""

    def _fmt_date_official(d):
        """Official template uses DD-Mon-YY format (e.g. '01-Apr-26')."""
        if d is None:
            return ""
        try:
            return d.strftime("%d-%b-%y")
        except Exception:
            return str(d)

    def _state_label(state_code: str) -> str:
        """Convert '29' to '29-Karnataka' style."""
        STATE_NAMES = {
            "01": "Jammu and Kashmir", "02": "Himachal Pradesh", "03": "Punjab",
            "04": "Chandigarh", "05": "Uttarakhand", "06": "Haryana",
            "07": "Delhi", "08": "Rajasthan", "09": "Uttar Pradesh",
            "10": "Bihar", "11": "Sikkim", "12": "Arunachal Pradesh",
            "13": "Nagaland", "14": "Manipur", "15": "Mizoram",
            "16": "Tripura", "17": "Meghalaya", "18": "Assam",
            "19": "West Bengal", "20": "Jharkhand", "21": "Odisha",
            "22": "Chhattisgarh", "23": "Madhya Pradesh", "24": "Gujarat",
            "25": "Daman and Diu", "26": "Dadra and Nagar Haveli",
            "27": "Maharashtra", "28": "Andhra Pradesh", "29": "Karnataka",
            "30": "Goa", "31": "Lakshadweep", "32": "Kerala",
            "33": "Tamil Nadu", "34": "Puducherry",
            "35": "Andaman and Nicobar Islands", "36": "Telangana",
            "37": "Andhra Pradesh", "38": "Ladakh",
        }
        s = str(state_code or "").strip().zfill(2)
        if s in STATE_NAMES:
            return f"{s}-{STATE_NAMES[s]}"
        return s

    # ----- b2b,sez,de sheet -----
    b2b_invoices = buckets.get("b2b", [])
    if b2b_invoices:
        ws = wb.create_sheet("b2b,sez,de")
        headers = [
            "GSTIN/UIN of Recipient", "Receiver Name", "Invoice Number",
            "Invoice date", "Invoice Value", "Place Of Supply",
            "Reverse Charge", "Applicable % of Tax Rate", "Invoice Type",
            "E-Commerce GSTIN", "Rate", "Taxable Value", "Cess Amount",
        ]
        rows = []
        # Each invoice can have multiple rates → one row per rate within invoice
        for inv in b2b_invoices:
            # Group items by rate within this invoice
            from collections import defaultdict
            by_rate = defaultdict(lambda: {"txval": 0.0, "csamt": 0.0})
            for item in inv["items"]:
                rate = round(float(item["tax_rate"] or 0), 2)
                by_rate[rate]["txval"] += float(item["taxable_value"] or 0)
                by_rate[rate]["csamt"] += float(item["cess_amount"] or 0)
            inv_value = round(float(inv["invoice_value"]), 2)
            inv_date = _fmt_date_official(inv["invoice_date"])
            pos_label = _state_label(inv["place_of_supply"])
            for rate, agg in sorted(by_rate.items()):
                rows.append([
                    inv["gstin"],
                    inv["customer_name"],
                    str(inv["invoice_no"]).strip(),
                    inv_date,
                    inv_value,
                    pos_label,
                    "N",                       # Reverse Charge
                    "",                        # Applicable % of Tax Rate
                    "Regular B2B",             # Invoice Type
                    "",                        # E-Commerce GSTIN
                    rate,
                    round(agg["txval"], 2),
                    round(agg["csamt"], 2),
                ])
        _write_table(ws, 1, headers, rows)
        _autofit(ws)

    # ----- b2ba sheet (Amendments) — empty placeholder -----
    # B2BA is for amending PRIOR-PERIOD invoices. We don't currently support
    # amendment workflow, so we ship an empty sheet with the correct headers
    # so you can populate it manually if needed.
    ws = wb.create_sheet("b2ba")
    headers = [
        "GSTIN/UIN of Recipient", "Receiver Name", "Original Invoice Number",
        "Original Invoice date", "Revised Invoice Number", "Revised Invoice date",
        "Invoice Value", "Place Of Supply", "Reverse Charge",
        "Applicable % of Tax Rate", "Invoice Type", "E-Commerce GSTIN",
        "Rate", "Taxable Value", "Cess Amount",
    ]
    _write_table(ws, 1, headers, [])
    _autofit(ws)

    # ----- b2cs sheet -----
    b2cs_invoices = buckets.get("b2cs", [])
    if b2cs_invoices:
        ws = wb.create_sheet("b2cs")
        headers = [
            "Type", "Place Of Supply", "Rate", "Applicable % of Tax Rate",
            "Taxable Value", "Cess Amount", "E-Commerce GSTIN",
        ]
        # B2CS rolls up by (POS, rate, sply_ty)
        from collections import defaultdict
        consolidated = defaultdict(lambda: {"txval": 0.0, "csamt": 0.0})
        for inv in b2cs_invoices:
            for item in inv["items"]:
                pos = str(inv["place_of_supply"] or "").zfill(2)
                rate = round(float(item["tax_rate"] or 0), 2)
                key = (pos, rate)
                consolidated[key]["txval"] += float(item["taxable_value"] or 0)
                consolidated[key]["csamt"] += float(item["cess_amount"] or 0)
        rows = []
        for (pos, rate), agg in sorted(consolidated.items()):
            rows.append([
                "OE",                         # Type (OE = Other than E-commerce)
                _state_label(pos),
                rate,
                "",                           # Applicable % of Tax Rate
                round(agg["txval"], 2),
                round(agg["csamt"], 2),
                "",                           # E-Commerce GSTIN
            ])
        _write_table(ws, 1, headers, rows)
        _autofit(ws)

    # ----- cdnr sheet (Credit/Debit Notes — Registered) -----
    cdnr_notes = buckets.get("cdnr", [])
    if cdnr_notes:
        ws = wb.create_sheet("cdnr")
        headers = [
            "GSTIN/UIN of Recipient", "Receiver Name", "Note Number",
            "Note Date", "Note Type", "Place Of Supply", "Reverse Charge",
            "Note Supply Type", "Note Value", "Applicable % of Tax Rate",
            "Rate", "Taxable Value", "Cess Amount",
        ]
        rows = []
        from collections import defaultdict
        for note in cdnr_notes:
            by_rate = defaultdict(lambda: {"txval": 0.0, "csamt": 0.0})
            for item in note["items"]:
                rate = round(float(item["tax_rate"] or 0), 2)
                by_rate[rate]["txval"] += float(item["taxable_value"] or 0)
                by_rate[rate]["csamt"] += float(item["cess_amount"] or 0)
            note_value = round(float(note["invoice_value"]), 2)
            note_date = _fmt_date_official(note["invoice_date"])
            pos_label = _state_label(note["place_of_supply"])
            note_type = note.get("doc_type", "C")  # 'C' or 'D'
            for rate, agg in sorted(by_rate.items()):
                rows.append([
                    note["gstin"],
                    note["customer_name"],
                    str(note["invoice_no"]).strip(),
                    note_date,
                    note_type,
                    pos_label,
                    "N",
                    "Regular B2B",
                    note_value,
                    "",
                    rate,
                    round(agg["txval"], 2),
                    round(agg["csamt"], 2),
                ])
        _write_table(ws, 1, headers, rows)
        _autofit(ws)

    # ----- docs sheet (Documents Issued) -----
    regular_invoices = [d for d in invoices if d.get("doc_type", "INV") == "INV"]
    if regular_invoices:
        ws = wb.create_sheet("docs")
        headers = [
            "Nature of Document", "Sr. No. From", "Sr. No. To",
            "Total Number", "Cancelled",
        ]
        inv_nums = sorted([str(inv["invoice_no"]).strip()
                           for inv in regular_invoices if inv.get("invoice_no")])
        if inv_nums:
            rows = [[
                "Invoices for outward supply",
                inv_nums[0],
                inv_nums[-1],
                len(inv_nums),
                0,
            ]]
        else:
            rows = []
        _write_table(ws, 1, headers, rows)
        _autofit(ws)

    # ----- hsn(b2b) sheet — exact official template format -----
    if b2b_invoices:
        ws = wb.create_sheet("hsn(b2b)")
        _write_hsn_official(ws, b2b_invoices)

    # ----- hsn(b2c) sheet -----
    b2c_all = buckets.get("b2cl", []) + b2cs_invoices
    if b2c_all:
        ws = wb.create_sheet("hsn(b2c)")
        _write_hsn_official(ws, b2c_all)


def _write_hsn_official(ws, invoices_subset):
    """Write the HSN section in the EXACT column order of the official template:
    HSN, Description, UQC, Total Quantity, Total Value, Taxable Value,
    Integrated Tax Amount, Central Tax Amount, State/UT Tax Amount, Cess Amount, Rate
    """
    from collections import defaultdict

    UQC_LABELS = {
        "BAG": "BAG-BAGS", "BAL": "BAL-BALE", "BDL": "BDL-BUNDLES",
        "BOX": "BOX-BOX", "BTL": "BTL-BOTTLES", "BUN": "BUN-BUNCHES",
        "CAN": "CAN-CANS", "CTN": "CTN-CARTONS", "DOZ": "DOZ-DOZENS",
        "GMS": "GMS-GRAMMES", "KGS": "KGS-KILOGRAMS", "LTR": "LTR-LITRES",
        "MTR": "MTR-METERS", "NOS": "NOS-NUMBERS", "PAC": "PAC-PACKS",
        "PCS": "PCS-PIECES", "PRS": "PRS-PAIRS", "ROL": "ROL-ROLLS",
        "SET": "SET-SETS", "TON": "TON-TONNES", "UNT": "UNT-UNITS",
        "OTH": "OTH-OTHERS", "NA": "NA",
    }
    def _uqc(u):
        s = str(u or "").upper().strip().rstrip(".")
        if s in UQC_LABELS:
            return UQC_LABELS[s]
        if s in ("NO", "NOS.", "NUMBER", "NUMBERS"):
            return UQC_LABELS["NOS"]
        if s in ("KG", "KILOGRAM", "KILOGRAMS"):
            return UQC_LABELS["KGS"]
        if s in ("PC", "PIECE", "PIECES"):
            return UQC_LABELS["PCS"]
        return UQC_LABELS["OTH"]

    groups = defaultdict(lambda: {
        "qty": 0.0, "txval": 0.0,
        "iamt": 0.0, "camt": 0.0, "samt": 0.0, "csamt": 0.0,
        "uqc_raw": "",
    })
    for inv in invoices_subset:
        for item in inv["items"]:
            hsn = str(item.get("hsn", "")).strip()
            rate = round(float(item.get("tax_rate", 0)), 2)
            key = (hsn, rate)
            g = groups[key]
            # Excel template keeps actual UQC and quantity even for services
            # (unlike the JSON which uses NA/0 for services).
            g["qty"] += float(item.get("quantity", 0) or 0)
            if not g["uqc_raw"]:
                g["uqc_raw"] = item.get("uqc", "OTH")
            g["txval"] += float(item.get("taxable_value", 0) or 0)
            g["iamt"] += float(item.get("igst_amount", 0) or 0)
            g["camt"] += float(item.get("cgst_amount", 0) or 0)
            g["samt"] += float(item.get("sgst_amount", 0) or 0)
            g["csamt"] += float(item.get("cess_amount", 0) or 0)

    headers = [
        "HSN", "Description", "UQC", "Total Quantity", "Total Value",
        "Taxable Value", "Integrated Tax Amount", "Central Tax Amount",
        "State/UT Tax Amount", "Cess Amount", "Rate",
    ]
    rows = []
    for (hsn, rate), g in sorted(groups.items()):
        total = g["txval"] + g["iamt"] + g["camt"] + g["samt"] + g["csamt"]
        rows.append([
            hsn, "",
            _uqc(g["uqc_raw"]),
            round(g["qty"], 2),
            round(total, 2),
            round(g["txval"], 2),
            round(g["iamt"], 2),
            round(g["camt"], 2),
            round(g["samt"], 2),
            round(g["csamt"], 2),
            rate,
        ])
    _write_table(ws, 1, headers, rows)
    _autofit(ws)
